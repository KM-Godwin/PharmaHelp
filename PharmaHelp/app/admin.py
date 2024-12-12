from django.contrib import admin
from django.http import HttpResponse, JsonResponse
from django.utils.html import format_html
from django.db.models import Sum, Count, F, Q, Avg
from django.db.models.functions import TruncMonth, TruncDay, ExtractMonth
from django.urls import path
from django.template.response import TemplateResponse
from django.utils import timezone
from datetime import datetime, timedelta
import csv
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from decimal import Decimal
from .models import Drug, Invoice, StockMovement, Sale

class AdvancedExportMixin:
    def get_field_value(self, obj, field):
        """Get the value of a field, handling related fields and callables"""
        attrs = field.split('__')
        value = obj
        for attr in attrs:
            if hasattr(value, attr):
                value = getattr(value, attr)
                if callable(value):
                    value = value()
        return value

    def export_as_csv(self, modeladmin, request, queryset):
        meta = self.model._meta
        field_names = self.get_export_fields()

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename={meta.verbose_name_plural.lower()}.csv'
        writer = csv.writer(response)

        writer.writerow(self.get_export_headers())
        for obj in queryset:
            row = [self.get_field_value(obj, field) for field in field_names]
            writer.writerow(row)

        return response
    export_as_csv.short_description = "Export Selected to CSV"

    def export_as_excel(self, modeladmin, request, queryset):
        meta = self.model._meta
        field_names = self.get_export_fields()

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = meta.verbose_name_plural

        # Style for headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Write headers
        headers = self.get_export_headers()
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Write data
        for row, obj in enumerate(queryset, 2):
            for col, field in enumerate(field_names, 1):
                value = self.get_field_value(obj, field)
                worksheet.cell(row=row, column=col, value=str(value))

        # Adjust column widths
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)

        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={meta.verbose_name_plural.lower()}.xlsx'
        return response
    export_as_excel.short_description = "Export Selected to Excel"

    def export_as_json(self, modeladmin, request, queryset):
        meta = self.model._meta
        field_names = self.get_export_fields()

        data = []
        for obj in queryset:
            item = {}
            for field in field_names:
                item[field] = str(self.get_field_value(obj, field))
            data.append(item)

        response = HttpResponse(
            json.dumps(data, indent=2),
            content_type='application/json'
        )
        response['Content-Disposition'] = f'attachment; filename={meta.verbose_name_plural.lower()}.json'
        return response
    export_as_json.short_description = "Export Selected to JSON"

    def get_actions(self, request):
        actions = super().get_actions(request)
        actions['export_as_csv'] = (
            self.export_as_csv,
            'export_as_csv',
            "Export Selected to CSV"
        )
        actions['export_as_excel'] = (
            self.export_as_excel,
            'export_as_excel',
            "Export Selected to Excel"
        )
        actions['export_as_json'] = (
            self.export_as_json,
            'export_as_json',
            "Export Selected to JSON"
        )
        return actions



@admin.register(Drug)
class DrugAdmin(AdvancedExportMixin, admin.ModelAdmin):
    list_display = ('name', 'barcode', 'category', 'in_Stock', 'stock_status', 
                   'price', 'expiry_Date', 'value_in_stock')
    list_filter = ('category', 'expiry_Date')
    search_fields = ('name', 'barcode')
    readonly_fields = ('created_at', 'updated_at', 'value_in_stock')

    def stock_status(self, obj):
        try:
            if obj.in_Stock <= obj.reorder_level:
                return format_html('<span style="color: red;">Low Stock</span>')
            elif obj.in_Stock <= (obj.reorder_level * Decimal('1.5')):
                return format_html('<span style="color: orange;">Medium Stock</span>')
            return format_html('<span style="color: green;">In Stock</span>')
        except (AttributeError, TypeError):
            return format_html('<span style="color: gray;">Not Set</span>')
    stock_status.short_description = 'Stock Status'

    def value_in_stock(self, obj):
        try:
            value = obj.in_Stock * obj.price
            return format_html('<span>{}</span>', '{:.2f}'.format(value))
        except (AttributeError, TypeError):
            return format_html('<span>0.00</span>')
    value_in_stock.short_description = 'Value in Stock'

    def get_export_fields(self):
        return ['name', 'barcode', 'category', 'in_Stock', 'price', 'expiry_Date', 'reorder_level']

    def get_export_headers(self):
        return ['Name', 'Barcode', 'Category', 'Stock', 'Price', 'Expiry Date', 'Reorder Level']




@admin.register(StockMovement)
class StockMovementAdmin(AdvancedExportMixin, admin.ModelAdmin):
    list_display = ('drug', 'quantity_changed', 'movement_type', 'date', 
                   'reason', 'reference_number')
    list_filter = ('movement_type', 'date', 'drug__category')
    search_fields = ('drug__name', 'reason', 'reference_number')
    date_hierarchy = 'date'

    def get_export_fields(self):
        return ['drug', 'quantity_changed', 'movement_type', 'date', 'reason', 
                'reference_number']

    def get_export_headers(self):
        return ['Drug', 'Quantity', 'Movement Type', 'Date', 'Reason', 
                'Reference Number']

@admin.register(Invoice)
class InvoiceAdmin(AdvancedExportMixin, admin.ModelAdmin):
    list_display = ('invoice_number', 'drug', 'quantity', 'unit_Price', 
                   'formatted_total_Price', 'invoice_date', 'payment_status', 
                   'payment_badge')
    list_filter = ('payment_status', 'invoice_date', 'drug__category')
    search_fields = ('drug__name', 'invoice_number')
    date_hierarchy = 'invoice_date'

    def payment_badge(self, obj):
        colors = {
            'PAID': 'green',
            'PENDING': 'orange',
            'OVERDUE': 'red'
        }
        color = colors.get(obj.payment_status, 'gray')
        return format_html('<span style="color: {};">{}</span>', 
                         color, obj.get_payment_status_display())
    payment_badge.short_description = 'Payment Status'

    def get_export_fields(self):
        return ['invoice_number', 'drug', 'quantity', 'unit_Price', 'total_Price', 
                'invoice_date', 'payment_status']

    def get_export_headers(self):
        return ['Invoice Number', 'Drug', 'Quantity', 'Unit Price', 'Total Price', 
                'Invoice Date', 'Payment Status']


@admin.register(Sale)
class SaleAdmin(AdvancedExportMixin, admin.ModelAdmin):
    list_display = ('drug', 'quantity', 'payment_method', 'total_amount',
                   'date', 'user')
    list_filter = ('payment_method', 'date', 'drug__category')
    search_fields = ('drug__name', 'user__username')
    date_hierarchy = 'date'

    def get_export_fields(self):
        return ['drug', 'quantity', 'payment_method', 'total_amount', 'date',
                'user']

    def get_export_headers(self):
        return ['Drug', 'Quantity', 'Payment Method', 'Total Amount', 'Date',
                'User']